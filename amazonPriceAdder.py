from selenium import webdriver
from time import sleep
from openpyxl import Workbook

wb = Workbook()
ws = wb.active


def checkNameInSheet(name):
    global ws
    for x in range(2, getNextEmptyRow(1)):
        if (ws.cell(row=x, column=1).value == name):
            return True
        if (ws.cell(row=x, column=1).value == None):
            break;

    return False

def findNameInSheet(name):
    global ws
    r = 2
    for x in range(2, getNextEmptyRow(1)):
        if (ws.cell(row=x, column=1).value == name):
            return r
        r+=1
    return -1

def getNextEmptyColumn(r):
    global  ws
    nonEmpty = 0
    i = 1
    while (not ws.cell(row=r,column=i).value == None):
        nonEmpty+=1
        i+=1
    return nonEmpty + 1

def getNextEmptyRow(column):
    return getNumberOfRows(column) + 2


def getNumberOfRows(col): #Does not include the first row containing the titles.
    global  ws
    nonEmpty = 0
    i = 2
    while (not ws.cell(row=i,column=col).value == None):
        nonEmpty+=1
        i+=1
    return nonEmpty



username = "username"
password = "password"

browser = webdriver.Chrome()
browser.get("https://www.amazon.com/ap/signin?_encoding=UTF8&ignoreAuthState=1&openid.assoc_handle=usflex&openid.claimed_id=http%3A%2F%2Fspecs.openid.net%2Fauth%2F2.0%2Fidentifier_select&openid.identity=http%3A%2F%2Fspecs.openid.net%2Fauth%2F2.0%2Fidentifier_select&openid.mode=checkid_setup&openid.ns=http%3A%2F%2Fspecs.openid.net%2Fauth%2F2.0&openid.ns.pape=http%3A%2F%2Fspecs.openid.net%2Fextensions%2Fpape%2F1.0&openid.pape.max_auth_age=0&openid.return_to=https%3A%2F%2Fwww.amazon.com%2Fyour-account%3Fref_%3Dnav_signin&switch_account=")
browser.find_element_by_name("email").send_keys(username)
sleep(1)
browser.find_element_by_id("continue").click()
sleep(1)
browser.find_element_by_name("password").send_keys(password)
sleep(1)
browser.find_element_by_id("signInSubmit").click()
sleep(1)
browser.get("https://www.amazon.com/gp/your-account/order-history?ref_=ya_d_c_yo")
sleep(1)
browser.find_element_by_class_name("a-dropdown-container").click()
sleep(1)
li = browser.find_elements_by_class_name("a-dropdown-link")




#Load the years in to the spreadsheet
years = []
for year in li[2:]:
    years.append(year.text)

for x in range(len(years)):
    ws.cell(row=1,column=x+2).value = years[x]

wb.save("amazonResults.xlsx")

for year in years:  #loop for the years
    browser.get("https://www.amazon.com/gp/your-account/order-history?opt=ab&digitalOrders=1&unifiedOrders=1&returnTo=&orderFilter=year-" + str(year) )
    sleep(2)
    orderNum = browser.find_element_by_class_name("num-orders").text
    orderNum = int(orderNum.split()[0])

    pages = int(orderNum // 10 + 1)

    for x in range(pages):  #loop for the pages
        browser.get("https://www.amazon.com/gp/your-account/order-history/ref=oh_aui_pagination_1_2?ie=UTF8&orderFilter=year-" + str(year) + "&search=&startIndex=" + str(x * 10) )
        sleep(1)
        orders = browser.find_elements_by_class_name("value")        #Need the elements at 1(price) and 2(person) and then add 4

        priceLoop = range(1,len(orders),4)
        nameLoop = range(2,len(orders),4)

        prices = []
        names = []

        for p in priceLoop:
            prices.append(float(orders[p].text[1:]))
        for n in nameLoop:
            names.append(orders[n].text)

        for i in range(len(prices)):
            if checkNameInSheet(names[i]):
                r = findNameInSheet(names[i])
                c = getNextEmptyColumn(r)
                ws.cell(row=r,column=c).value = prices[i]
            else:
                r = getNextEmptyRow(1)
                ws.cell(row = r, column = 1).value = names[i]
                ws.cell(row=r,column =2).value = prices[i]
        wb.save("amazonResults.xlsx")
        sleep(1)







def checkNameInSheet(name):
    global ws
    for x in range(2, getNextEmptyRow(1)):
        if (ws1.cell(row=x, column=1).value == name):
            return True
        if (ws1.cell(row=x, column=1).value == None):
            break;

    return False

def findNameInSheet(name):
    global ws
    r = 2
    for x in range(2, getNextEmptyRow(1)):
        if (ws1.cell(row=x, column=1).value == name):
            return r
        r+=1
    return -1

def getNextEmptyColumn(r):
    global  ws
    nonEmpty = 0
    i = 1
    while (not ws.cell(row=r,column=i).value == None):
        nonEmpty+=1
        i+=1
    return nonEmpty + 1

def getNextEmptyRow(column):
    return getNumberOfRows(column) + 2


def getNumberOfRows(col): #Does not include the first row containing the titles.
    global  ws
    nonEmpty = 0
    i = 2
    while (not ws.cell(row=i,column=col).value == None):
        nonEmpty+=1
        i+=1
    return nonEmpty